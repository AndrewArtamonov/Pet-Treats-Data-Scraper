import re
import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement

from libraries.common import get_report_file_headers


class PetLovers:
    def __init__(self):
        # self.dry_food_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/dry-food?SortBy=Brand&SortOrder=Asc"
        self.dry_food_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/dry-food?GroupID=&PageNum=12&SortBy=Brand&SortOrder=Asc"
        self.food_pouches_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/food-pouches?SortBy=Brand&SortOrder=Asc"
        self.canned_food_url = "https://www.petloverscentre.com/dog/dog-food-treats/food/canned-food?SortBy=Brand&SortOrder=Asc"
        self.broth_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/broth?SortBy=Brand&SortOrder=Asc"
        self.air_dried_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/air-dried?SortBy=Brand&SortOrder=Asc"
        self.dehydrated_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/dehydrated?SortBy=Brand&SortOrder=Asc"
        self.freeze_dried_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/freeze-dried?SortBy=Brand&SortOrder=Asc"
        self.frozen_url = "https://www.petloverscentre.com/dog/dog-food-treats/dog-dog-food-treats-food/frozen-food?SortBy=Brand&SortOrder=Asc"

        self.chromedriver_path = "/Users/andriyartamonov/Documents/chromedriver"
        options = Options()
        options.headless = True
        self.browser = webdriver.Chrome(executable_path=self.chromedriver_path, options=options)
        self.browser.set_page_load_timeout(300)
        # self.browser = webdriver.Chrome(executable_path=self.chromedriver_path)

        self.report_file_path = "Scraping Masterlist.xlsx"
        self.report_file = load_workbook(filename=self.report_file_path, read_only=False)
        self.pet_lovers_sheet = self.report_file["Pet Lovers Centre"]
        self.report_file_headers = get_report_file_headers(self.pet_lovers_sheet)

        self.page = 1
        # self.products_processed = 2  # FOR NEW
        self.products_processed = 665
        self.nutrition = []

        self.start_time = time.time()

    def process(self):
        # url = ""  #
        # self.process_product("", 2, url)

        # Dry Food
        self.process_base(product_type="Dry Food", product_class="Dry Food",
                          url=self.dry_food_url, comment="Opening Dry Food main page")

        # Canned / Moist food
        self.process_base(product_type="Canned/Moist", product_class="Food Pouches",
                          url=self.food_pouches_url, comment="Opening Food Pouches main page")
        self.process_base(product_type="Canned/Moist", product_class="Canned Food",
                          url=self.canned_food_url, comment="Opening Canned Food main page")
        self.process_base(product_type="Canned/Moist", product_class="Broth",
                          url=self.broth_url, comment="Opening Broth Food main page")
        print("-"*40)
        print("COMPLETED Canned / Moist food GROUP")
        print("-"*40)

        # Dehydrated/Freeze Dried food
        self.process_base(product_type="Dehydrated/Freeze Dried", product_class="Air-Dried",
                          url=self.air_dried_url, comment="Opening Air Dried Food main page")
        self.process_base(product_type="Dehydrated/Freeze Dried", product_class="Dehydrated",
                          url=self.dehydrated_url, comment="Opening Dehydrated Food main page")
        self.process_base(product_type="Dehydrated/Freeze Dried", product_class="Freeze-Dried",
                          url=self.freeze_dried_url, comment="Opening Freeze Dried Food main page")
        print("-"*40)
        print("COMPLETED Dehydrated/Freeze Dried food GROUP")
        print("-"*40)

        # Frozen/Raw
        self.process_base(product_type="Frozen/Raw", product_class="Frozen food",
                          url=self.frozen_url, comment="Opening Frozen Food main page")

    def process_base(self, product_type, product_class, url, comment):
        print("*" * 30)
        print(comment)
        self.browser.get(url=url)
        self.browser.maximize_window()

        self.process_pages(product_class, product_type)

    def process_pages(self, product_class, product_type):
        # for _ in range(4):
        while True:
            print(f"Processing {self.page} page")
            product_xpath = "//div[@class='product-box col-1-3 tab-col-1-2 mobile-col-1-1']/div[@class='prod-img']//a"
            page_products_links: list = self.browser.find_elements(by=By.XPATH, value=product_xpath)

            for idx, product in enumerate(page_products_links, start=2):
                self.process_product(product, idx, product_class, product_type)

            print(f"Time spended for page - {round((time.time() - self.start_time) / 60, 1)} minutes")
            self.start_time = time.time()

            next_page_button_xpath = "//div[@class='page-no-top col-2-3']/ul[@class='page-prev-next']"
            button_element = self.browser.find_element(by=By.XPATH, value=next_page_button_xpath)
            button_size = button_element.size['width']
            if button_size:
                self.page += 1
                button_element.click()
            else:
                print("ALL PAGES COMPLETED")
                return

    def process_product(self, product, row_idx, product_class, product_type, product_url=None):
        product_url: str = product.get_attribute('href') if not product_url else product_url
        print(f"Processing {row_idx - 1} product with url - {product_url}")
        if self.pet_lovers_sheet[f"c{self.products_processed}"].value == product_url:
            print("WAS ALREADY PROCESSED")
            self.products_processed += 1
            return

        self.browser.switch_to.new_window()
        self.browser.get(url=product_url)

        # Get brand name
        brand_name_xpath = "//p[@class='small-name']"
        brand_name: str = self.browser.find_element(by=By.XPATH, value=brand_name_xpath).text

        # Get title text
        title_xpath = "//div[@class='prod-details-top col-1-1']/h1"
        title_text: str = self.browser.find_element(by=By.XPATH, value=title_xpath).text

        # Get product details text
        product_details_xpath = "//div[@id='details']"
        product_details_text: str = self.browser.find_element(by=By.XPATH, value=product_details_xpath).text

        # Get product data
        product_size, product_title = self.get_product_title_and_size(brand_name, title_text)
        product_image_url: str = self.get_product_image_url()
        product_price: str = self.browser.find_element(by=By.XPATH, value="//div[@class='prod-price']/p[1]").text
        product_size: str = self.get_product_size(product_details_text,
                                                  title_text) if not product_size else product_size.lower()
        product_elements = self.get_product_elements(product_details_text)
        product_origin = self.get_product_origin(product_details_text)
        product_ingredients = self.get_product_ingredients(product_details_text)

        product_data = ["Pet Lovers Centre", product_title, product_url, product_image_url, product_type, product_class,
                        product_size, product_price, product_origin, product_ingredients]

        # Print product main data
        for column_idx, value in enumerate(product_data, start=1):
            self.pet_lovers_sheet.cell(self.products_processed, column_idx, value)

        # Print product nutrition data
        for key, value in product_elements.items():
            if key not in self.report_file_headers:
                new_header_index = len(self.report_file_headers) + 1
                self.pet_lovers_sheet.cell(1, new_header_index, key)
                self.pet_lovers_sheet.cell(self.products_processed, new_header_index, value)
                self.report_file_headers.append(key)
            else:
                column_idx = self.report_file_headers.index(key) + 1
                self.pet_lovers_sheet.cell(self.products_processed, column_idx, value)

        self.products_processed += 1
        self.report_file.save(filename=self.report_file_path)
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[0])

    @staticmethod
    def get_product_title_and_size(brand_name, title_text):
        product_size = ""
        product_size_in_title = re.findall(r"-? ?(\d+(?:\.\d+)? ?(?:g|kgs?|lbs|oz))", title_text, re.IGNORECASE)
        if "lbs" in title_text and "kg" in title_text:
            regexp_size_pattern = re.compile(r" \d+(?:\.\d+)?lbs ?\((\d+(?:\.\d+)?kg)\)")
            product_size = re.findall(pattern=regexp_size_pattern, string=title_text)[0]
            product_title = brand_name + " - " + re.sub(pattern=regexp_size_pattern, repl="", string=title_text)
        elif product_size_in_title:
            product_size = product_size_in_title[0]
            product_title = title_text.replace(product_size, "").strip()
            product_title: str = brand_name + " - " + product_title
        else:
            product_title: str = brand_name + " - " + title_text
        return product_size, product_title

    @staticmethod
    def get_product_ingredients(product_details_text):
        product_ingredients_regexp = r"Ingredients:? ?\n([\s\S]+?)(?:Analysis|Made in|Size)"
        product_ingredients: str = re.findall(pattern=product_ingredients_regexp, string=product_details_text)[0]
        product_ingredients = product_ingredients.replace("\n", "").replace(".SUPPLEMENTS:", ",").replace(
            ".Supplements:", "")
        if not product_ingredients:
            print("ERROR NO PRODUCT INGREDIENTS")
        return product_ingredients

    @staticmethod
    def get_product_origin(product_details_text):
        product_origin: list = re.findall(pattern=r"Made in (?:the )?(.+)", string=product_details_text)
        product_origin: str = product_origin[0] if product_origin else ""
        if not product_origin:
            print("ERROR NO COUNTRY ORIGIN")
        return product_origin

    def get_product_elements(self, product_details_text):
        product_elements = {}
        if "analysis" in product_details_text.lower():
            product_nutrition_table_xpath = "(//div[@id='details']//table)[1]//tr"
            product_nutrition_table = self.browser.find_elements(by=By.XPATH, value=product_nutrition_table_xpath)
            for element in product_nutrition_table:
                name_value_separator_xpath = "./td"
                row_cells = element.find_elements(by=By.XPATH, value=name_value_separator_xpath)
                if len(row_cells) == 2:
                    name, value = row_cells
                else:
                    print("ERROR ANALYSIS TABLE IS NOT CORRECT!!!")
                    return product_elements

                name = self.process_nutrition_name(name)
                value = self.process_nutrition_value(value)

                product_elements[name] = value
            if len(product_elements.keys()) != len(product_nutrition_table):
                print("ERROR DURING NUTRITION PARSING")
            if not len(product_elements.keys()):
                print("ERROR ANALYSIS NOT FOUND!!!")
        else:
            print("ERROR ANALYSIS NOT FOUND!!!")
        return product_elements

    def get_product_image_url(self):
        image_div_xpath = "//div[@class='zoomContainer']/div/div"
        image_div_element: WebElement = self.browser.find_element(by=By.XPATH, value=image_div_xpath)
        image_div_style: str = image_div_element.get_attribute("style")
        product_image_url: str = re.findall(pattern=r'url\("(.+)"\)', string=image_div_style)[0]
        return product_image_url

    @staticmethod
    def get_product_size(product_details_text: str, title_text: str):
        product_size_from_title = re.findall(pattern=r"(\d+(?:\.\d+)? ?kg)", string=title_text)
        product_size = product_size_from_title[0] if product_size_from_title else ""
        if not product_size:
            product_size_from_description: list = re.findall(pattern=r"Size:\n(.+)", string=product_details_text)
            product_size = product_size_from_description[0] if product_size_from_description else ""
        product_size = product_size.strip().replace(" ", "")
        return product_size

    @staticmethod
    def process_nutrition_name(name: WebElement) -> str:
        name = name.text.strip().replace("*", "")
        name = "Arginine" if "arginine" in name.lower() else name
        name = "Ascorbic Acid" if "ascorbic acid" in name.lower() else name
        name = "Beta-Carotene" if "beta-carotene" in name.lower() else name
        name = "Glucosamine" if "glucosamine" in name.lower() else name
        name = "Chondroitin" if "chondroitin" in name.lower() else name
        name = "Chloride" if "chloride" in name.lower() else name
        name = "Phosphorus" if "phosphorus" in name.lower() else name
        name = "Phosphorus" if "phosphorous" in name.lower() else name
        name = "Phosphorus" if "phosporous" in name.lower() else name
        name = "Potassium" if "potassium" in name.lower() else name
        name = "Phosphorous" if "phosphorous" in name.lower() or "phosphorus" in name.lower() else name
        name = "Crude Ash" if "ash" in name.lower() else name
        name = "Crude Protein" if "protein" in name.lower() else name
        name = "Crude Fat" if name == "Fats" else name
        name = "Crude Fat" if name == "Crude Oils and Fats" else name
        name = "Crude Fat" if name == "Crude oil fats" else name
        name = "Crude Fat" if "fat" in name.lower() and not re.findall(pattern=r"fat\w+", string=name,
                                                                       flags=re.IGNORECASE) else name
        name = "Crude Fiber" if "fiber" in name.lower() or "fibres" in name.lower() or "fibre" in name.lower() else name
        name = "DHA" if "dha" in name.lower() else name
        name = "EPA" if "epa" in name.lower() else name
        name = "NFE" if "nfe" in name.lower() else name
        name = "Calories" if "calori" in name.lower() else name
        name = "Calcium" if "calcium" in name.lower() else name
        name = "Cellulase" if "cellulase" in name.lower() else name
        name = "Copper" if "copper" in name.lower() else name
        name = "Linoleic Acid" if "linoleic acid" in name.lower() else name
        name = "Linoleic Acid" if "linoleicc acid" in name.lower() else name
        name = "L-Carnitine" if "l carnitine" in name.lower() else name
        name = "L-Carnitine" if "l-carnitine" in name.lower() else name
        name = "Lysine" if "lysine" in name.lower() else name
        name = "Iron" if "iron" in name.lower() else name
        name = "Taurin" if "taurin" in name.lower() else name
        name = "Tryptophan" if "tryptophan" in name.lower() else name
        name = "Chicory Root" if "chicory root" in name.lower() else name
        name = "Copper" if "coppper" in name.lower() else name
        name = "Lactobacillus Acidophilus" if "lactobacillus acidophilus" in name.lower() else name
        name = "Linoleic Acid" if "linoleic acid" in name.lower() else name
        name = "Magnesium" if "magnesium" in name.lower() else name
        name = "Moisture" if "moisture" in name.lower() else name
        name = "Crude Oils and Fats" if "crude oils and fats" in name.lower() else name
        name = "Metabolisable Energy" if re.findall(pattern=r"^ME:? ?", string=name) else name
        name = "Metabolisable Energy" if "metabolisable energy" in name.lower() else name
        name = "Metabolisable Energy" if "metabolizable energy" in name.lower() else name
        name = "Metabolisable Energy" if "metabolized energy" in name.lower() else name
        name = "Methionine" if "methionine" in name.lower() else name
        name = "Sodium" if "sodium" in name.lower() else name
        name = "Total Microorganisms" if "total microorganisms" in name.lower() else name
        name = "Total Microorganisms" if "total micro-organisms" in name.lower() else name
        name = "Zinc" if "zinc" in name.lower() else name
        name = "Omega-3" if re.findall(pattern=r"omega[-‑ ]?3", string=name, flags=re.IGNORECASE) else name
        name = "Omega-6" if re.findall(pattern=r"omega[-‑ ]?6", string=name, flags=re.IGNORECASE) else name
        name = "Omega-7" if re.findall(pattern=r"omega[-‑ ]?7", string=name, flags=re.IGNORECASE) else name
        name = "Omega-9" if re.findall(pattern=r"omega[-‑ ]?9", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin A" if re.findall(pattern=r"vitamin\.? a(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin B1" if re.findall(pattern=r"vitamin\.? b1(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin B2" if re.findall(pattern=r"vitamin\.? b2(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin B3" if re.findall(pattern=r"vitamin\.? b3(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin B5" if re.findall(pattern=r"vitamin\.? b5(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin B6" if re.findall(pattern=r"vitamin\.? b6(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin C" if re.findall(pattern=r"vitamin\.? c(?:$| |\))", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin D3" if re.findall(pattern=r"vitamin\.? d3(?:$| )", string=name, flags=re.IGNORECASE) else name
        name = "Vitamin E" if re.findall(pattern=r"vitamin\.? e(?:$| )", string=name, flags=re.IGNORECASE) else name
        return name

    @staticmethod
    def process_nutrition_value(value: WebElement) -> str:
        if "min" in value.text.lower() and "max" in value.text.lower():
            value = re.sub(pattern=r"min.+,", repl="", string=value.text, flags=re.IGNORECASE)
        else:
            value = value.text
        value = value.replace("minimum", "").replace("maximum", "")
        value = value.replace("Min.", "").replace("Max.", "")
        value = value.replace("min.", "").replace("max.", "")
        value = value.replace("Min", "").replace("Max", "")
        value = value.replace("min", "").replace("max", "").replace("()", "").replace("-", "")
        if 'IU/kg' in value:
            value = re.sub(r"(\d) (\d)", r"\1\2", value)
            value = value.replace(",", "")
        elif "%" in value:
            value = value.replace(" ", "")
        value = re.sub(r"\.00%", ".0%", value)
        value = value.strip()
        return value
